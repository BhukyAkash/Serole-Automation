from openpyxl import load_workbook
import os

CURRENT_DIR = os.path.dirname(__file__)
EXCEL_PATH = os.path.join(CURRENT_DIR, "..", "SIT", "Test Data.xlsx")

# ==== Input: Motor Test Data ======
VEHICLE_CONFIG = {
    "CV": {"reg_col": 11,  "ic_col": 12,  "used_col": 13},   # A, B, D
    "PC": {"reg_col": 6,  "ic_col": 7,  "used_col": 8},   # F, G, I
    "MC": {"reg_col": 1, "ic_col": 2, "used_col": 4},  # K, L, N
}

def get_vehicle_data(vehicle_type):
    cfg = VEHICLE_CONFIG[vehicle_type]
    wb = load_workbook(EXCEL_PATH)
    sheet = wb["Test Data"]

    claimed_row = None

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        used_val = sheet.cell(row=row_idx, column=cfg["used_col"]).value
        used_str = str(used_val).strip().upper() if used_val else ""

        if used_str in ("Y", "RUNNING"):
            continue

        reg   = sheet.cell(row=row_idx, column=cfg["reg_col"]).value
        mykad = sheet.cell(row=row_idx, column=cfg["ic_col"]).value

        if not reg or not mykad:
            break  # Empty row — no more test data

        sheet.cell(row=row_idx, column=cfg["used_col"]).value = "RUNNING"
        wb.save(EXCEL_PATH)
        claimed_row = row_idx
        break

    if claimed_row is None:
        print(f"No test data found for {vehicle_type}")
        return None

    return {
        "vehicle_reg_no": reg,
        "mykad": mykad,
        "claimed_row": claimed_row,
        "vehicle_type": vehicle_type,
    }

def mark_policy_issued(vehicle_type, claimed_row):
    """Call this after policy is successfully issued."""
    cfg = VEHICLE_CONFIG[vehicle_type]
    wb = load_workbook(EXCEL_PATH)
    sheet = wb["Test Data"]
    sheet.cell(row=claimed_row, column=cfg["used_col"]).value = "Y"
    wb.save(EXCEL_PATH)


def reset_on_error(vehicle_type, claimed_row):
    """Call this in except block if policy was NOT issued — returns row to pool."""
    cfg = VEHICLE_CONFIG[vehicle_type]
    wb = load_workbook(EXCEL_PATH)
    sheet = wb["Test Data"]
    sheet.cell(row=claimed_row, column=cfg["used_col"]).value = "N"
    wb.save(EXCEL_PATH)