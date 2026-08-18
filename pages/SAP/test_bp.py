import os, pytest
from openpyxl import load_workbook
from SAP.sap_utils import login

# pytest -s SAP\test_bp.py

def get_id_number(sheet_name):
    excel_path = os.path.join(os.path.dirname(__file__), "BP Test Data.xlsx")
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]

    row = 2
    while True:
        b_val = sheet.cell(row=row, column=2).value  # column B
        c_val = sheet.cell(row=row, column=3).value  # column C
        if b_val is None:
            raise ValueError(f"No more ID numbers found in '{sheet_name}' (row {row} is empty)")
        if c_val is None or str(c_val).strip() == "":
            break
        row += 1

    id_number = str(b_val).replace("-", "")
    return id_number, sheet, wb, excel_path, row

def save_bp(sheet, wb, excel_path, row, bp):
    sheet.cell(row=row, column=3).value = bp
    wb.save(excel_path)

# pytest -s SAP\test_bp.py::test_bp_pc
@pytest.mark.no_network_logger
def test_bp_pc(page):
    try:
        page.goto("https://tus4appuat.tuneprotect.com:44303/sap/bc/ui2/flp#BusinessPartner-maintain?sap-ui-tech-hint=GUI")
        login(page)

        frame = page.locator("iframe[title=\"Application\"]").content_frame

        id_number, sheet, wb, excel_path, row = get_id_number("PC")
        print("Mykad ID: ", id_number)
        frame.get_by_role("textbox", name="ID number").fill(id_number)
        frame.get_by_role("button", name="Start").click()

        partner = frame.locator("span[id='grid#C105#1,1#if']")
        page.wait_for_timeout(3000)
        if partner.is_visible():
            bp  = partner.inner_text().strip()
            print("Business Partner: ", bp)
            save_bp(sheet, wb, excel_path, row, bp)
        else:
            bp = "No BP Found"
            print("Business Partner:", bp)
            save_bp(sheet,wb,excel_path,row,bp)

    finally:
        page.locator("#meAreaHeaderButton").click()
        page.get_by_text("Sign Out").click()
        page.get_by_role("button", name="OK").click()

# pytest -s SAP\test_bp.py::test_bp_mc
@pytest.mark.no_network_logger
def test_bp_mc(page):
    try:
        page.goto("https://tus4appuat.tuneprotect.com:44303/sap/bc/ui2/flp#BusinessPartner-maintain?sap-ui-tech-hint=GUI")
        login(page)

        frame = page.locator("iframe[title=\"Application\"]").content_frame

        id_number, sheet, wb, excel_path, row = get_id_number("MC")
        print("Mykad ID: ", id_number)
        frame.get_by_role("textbox", name="ID number").fill(id_number)
        frame.get_by_role("button", name="Start").click()

        partner = frame.locator("span[id='grid#C105#1,1#if']")
        if partner.is_visible():
            bp  = partner.inner_text().strip()
            print("Business Partner: ", bp)
            save_bp(sheet, wb, excel_path, row, bp)
        else:
            bp = "No BP Found"
            print("Business Partner:", bp)
            save_bp(sheet,wb,excel_path,row,bp)

    finally:
        page.locator("#meAreaHeaderButton").click()
        page.get_by_text("Sign Out").click()
        page.get_by_role("button", name="OK").click()