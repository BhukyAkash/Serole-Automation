import os
from openpyxl import load_workbook

excel_path = os.path.join(os.path.dirname(__file__), "BP Test Data.xlsx")
wb = load_workbook(excel_path)

sheet_pc = wb["PC"]
sheet_mc = wb["MC"]

# --- PC Test Data ----
cell_pc = 26
# --- MC Test Data ---- 
cell_mc = 5

info = {
    "BP_PC" : sheet_pc.cell(row=cell_pc, column=3).value,
    "BP_MC" : sheet_mc.cell(row=cell_mc, column=3).value,
    "CC"    : "2210001629",
    "date"  : "01.06.2026",

    "MC" : {
        "pm_id"         : "MTPLMC000000",
        "vehicle_no"    : sheet_mc.cell(row=cell_mc, column=1).value,
        "coverage_type" : "Comprehensive",     # Third Party
        "covpac"        : "Comprehensive",       
        "si"            : "10000"
    },

    "PC" : {
        "pm_id"         : "MTPLPC000000",
        "vehicle_no"    : sheet_pc.cell(row=cell_pc, column=1).value,
        "si"            : "85000",  
        "coverage_type" : "Comprehensive",             # "Comprehensive"    "TP, Fire & Theft"
        "covpac"        : "Comprehensive",  # "Comprehensive"    "Third Party, Fire and Theft"
        }
}

def get_vehicle_info(vehicle_type: str) -> dict:
    if vehicle_type not in info:
        raise KeyError(
            f"Vehicle type '{vehicle_type}' not found. "
            f"Valid types: {list(info.keys())}"
        )
    return info[vehicle_type]
